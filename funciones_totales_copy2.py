# -*- coding: utf-8 -*-
"""
Created on Tue Nov 16 15:40:28 2021

@author: Fernando Isai Almaraz Fabian/Diego Rodriguez
Derechos reservados
"""

import pandas as pd
import numpy as np
import openpyxl

from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl import load_workbook
#

def renombrando_columna(df):
    
    lista_de_columna=list(df.columns)#Obteniendo la lista de columnas de una tabla.

    for i in range(len(lista_de_columna)):#Se empieza con un for porque repetiremos la accionn dentro del for en todas las columnas.
        j=0
        while df.iloc[j,i]=='0':#Importante
            j=j+1 
        else:
            if type (df.iloc[j,i])==str:
                if df.iloc[j,i][0:3]=='PR' or df.iloc[j,i][0:3]=='PT' or df.iloc[j,i][0:3]=='XX' or df.iloc[j,i][0:3]=='SD' or df.iloc[j,i][0:3]=='PP' or df.iloc[j,i][0:3]=='AV':
                    df=df.rename(columns={lista_de_columna[i]:'AVAA :Tyre INDST'})
                
                elif (df.iloc[j,i][4:5]=='/' and len(df.iloc[j,i])>=9) or (df.iloc[j,i][3:4]=='/' and len(df.iloc[j,i])>=9):
                    df=df.rename(columns={lista_de_columna[i]:'MISURA'})

                elif lista_de_columna[i]=='CPAA :ENCOLHIMENTO' or lista_de_columna[i]=='CPAA :ENCOLHIMENTO.1' or lista_de_columna[i]=='CPAA :ENCOLHIMENTO.2' or lista_de_columna[i]=='CPAA :ENCOLHIMENTO.3':
                    if df.iloc[j,i][0:3]=='SCP':
                        df=df.rename(columns={lista_de_columna[i]:'Complex SAP'})

                elif lista_de_columna[i]=='TEAB :A_SO_TEAB_DESCRIPT' or lista_de_columna[i]=='TEAB :A_SO_TEAB_DESCRIPT.1':
                    if all(x in list(df.iloc[:,i].unique()) for x in ['PARIAN', 'PALENQUE', 'ROMITA','REYNOSA','RIO','PEROTE'])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'TEAB :Description'})   

                elif lista_de_columna[i]== 'AVAA :ENCOLHIMENTO' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.1' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.2' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.3' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.4' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.5' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.6' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.7' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.8' or lista_de_columna[i]== 'AVAA :ENCOLHIMENTO.9': 
                    if df.iloc[j,i][0:2]=='XD' or df.iloc[j,i][0:2]=='RD':
                        df=df.rename(columns={lista_de_columna[i]:'AVAA :Sponge Code (L1)'})

                elif lista_de_columna[i]== 'CEBM :ENCOLHIMENTO' or lista_de_columna[i]=='CEBM :ENCOLHIMENTO.1':
                    if (df.iloc[j,i][0:2]=='CT'):
                        df=df.rename(columns={lista_de_columna[i]:'CEBM_COMPOUND_RC'})
                    elif (df.iloc[j,i][0:3]=='SCE'):
                        df=df.rename(columns={lista_de_columna[i]:'CEJA SAP Aro+filler'})

                elif lista_de_columna[i]=='FAAE :ENCOLHIMENTO' or lista_de_columna[i]=='FAAE :ENCOLHIMENTO.1' or lista_de_columna[i]=='FAAE :ENCOLHIMENTO.2' or lista_de_columna[i]=='FAAE :ENCOLHIMENTO.3':
                    if df.iloc[j,i][0:3]=='CDL' or df.iloc[j,i][0:3]=='CDR' or df.iloc[j,i][0:3]=='CDC' or df.iloc[j,i][0:3]=='CDW' or df.iloc[j,i][0:3]=='CDT' or df.iloc[j,i][0:3]=='CDO' :
                        df=df.rename(columns={lista_de_columna[i]:'DG Cap Compound'})
                    elif df.iloc[j,i][0:9]=='CELBA001C':
                        df=df.rename(columns={lista_de_columna[i]:'DL FAAE_COMPOUND_MF'})
                    elif df.iloc[j,i][0:9]=='CUNIT001C':
                        df=df.rename(columns={lista_de_columna[i]:'DJ Foglietta'})
                    elif df.iloc[j,i][0:3]=='SFA':
                        df=df.rename(columns={lista_de_columna[i]:'FAAE :ENCOLHIMENTO'})                    

                elif lista_de_columna[i]=='FNAB :ENCOLHIMENTO' or lista_de_columna[i]=='FNAB :ENCOLHIMENTO.1' or lista_de_columna[i]=='FNAB :ENCOLHIMENTO.2' or lista_de_columna[i]=='FNAB :ENCOLHIMENTO.3':
                    if df.iloc[j,i][0:5]=='CTURN' or df.iloc[j,i][0:4]=='CTLC':
                        df=df.rename(columns={lista_de_columna[i]:'FNAB_COMPOUND_AB'})
                    elif df.iloc[j,i][0:5]=='CEPRO' or df.iloc[j,i][0:5]=='CELBA' or df.iloc[j,i][0:5]=='CECOL':
                        df=df.rename(columns={lista_de_columna[i]:'FNAB_COMPOUND_FN'})                        
                    elif df.iloc[j,i][0:3]=='SFN':
                        df=df.rename(columns={lista_de_columna[i]:'FNAB :ENCOLHIMENTO'})
                        
                elif lista_de_columna[i]=='FAAP :ENCOLHIMENTO' or lista_de_columna[i]=='FAAP :ENCOLHIMENTO.1' or lista_de_columna[i]=='FAAP :ENCOLHIMENTO.2' or lista_de_columna[i]=='FAAP :ENCOLHIMENTO.3' or lista_de_columna[i]=='FAAP :ENCOLHIMENTO.4':
                    if df.iloc[j,i][0:5]=='CBANK' or df.iloc[j,i][0:5]=='CDLUM' or df.iloc[j,i][0:5]=='CDTCM' or df.iloc[j,i][0:5]=='CDFLI' or df.iloc[j,i][0:5]=='CDJPB' or df.iloc[j,i][0:5]=='CDOME' or df.iloc[j,i][0:5]=='CDFUM' or df.iloc[j,i][0:5]=='CDWEM' or df.iloc[j,i][0:5]=='CBIEN' or df.iloc[j,i][0:5]=='CBODA':
                        df=df.rename(columns={lista_de_columna[i]:'CV Cap Compound'})
                    elif df.iloc[j,i][0:9]=='CUNIT001C':
                        df=df.rename(columns={lista_de_columna[i]:'CY Foglietta'})                        
                    elif df.iloc[j,i][0:3]=='SFA':
                        df=df.rename(columns={lista_de_columna[i]:'FAAP :ENCOLHIMENTO'}) 
                    elif df.iloc[j,i][0:5]=='CELBA':
                        df=df.rename(columns={lista_de_columna[i]:'DA FAAP_COMPOUND_MF'})                         
                    elif all(x in list(df.iloc[:,i].unique()) for x in ['0','CUTE_001C','CUNIT001C','CUMAC001C','CUVOS001C'])==True or all(x in list(df.iloc[:,i].unique()) for x in ['0','CUNIT001C','CUMAC001C','CUVOS001C'])==True :
                        df=df.rename(columns={lista_de_columna[i]:'CW Base'})                       
                        

                elif lista_de_columna[i]=='FAAJ :ENCOLHIMENTO' or lista_de_columna[i]=='FAAJ :ENCOLHIMENTO.1' or lista_de_columna[i]=='FAAJ :ENCOLHIMENTO.2' or lista_de_columna[i]=='FAAJ :ENCOLHIMENTO.3' or lista_de_columna[i]=='FAAJ :ENCOLHIMENTO.4' or lista_de_columna[i]=='FAAJ :ENCOLHIMENTO.5' or lista_de_columna[i]=='FAAJ :ENCOLHIMENTO.6':      
                    if df.iloc[j,i][0:5]=='CDRIE' or df.iloc[j,i][0:5]=='CBERO' or df.iloc[j,i][0:5]=='CDASR' or df.iloc[j,i][0:5]=='CDLUM' or df.iloc[j,i][0:5]=='CDTCM' or df.iloc[j,i][0:5]=='CBODA' or df.iloc[j,i][0:5]=='CBANK' or df.iloc[j,i][0:5]=='CDFLI' or df.iloc[j,i][0:5]=='CDOME' or df.iloc[j,i][0:5]=='CDUFF' or df.iloc[j,i][0:5]=='CDJPB' or df.iloc[j,i][0:5]=='CDWEM' or df.iloc[j,i][0:5]=='CDFUM' or df.iloc[j,i][0:5]=='CDBIEN' or df.iloc[j,i][0:5]=='CBOOX' or df.iloc[j,i][0:4]=='C1VV' or df.iloc[j,i][0:5]=='CDJAE' or df.iloc[j,i][0:5]=='CDUMI' or df.iloc[j,i][0:6]=='CDIEGO' or df.iloc[j,i][0:5]=='CDFOG' or df.iloc[j,i][0:5]=='CDCAM' or df.iloc[j,i][0:4]=='C1RW' or df.iloc[j,i][0:4]=='C1VZ' or df.iloc[j,i][0:5]=='C1VW':
                        df=df.rename(columns={lista_de_columna[i]:'DR Cap Compound'})
                    elif all(x in list(df.iloc[:,i].unique()) for x in ['CUNIT001C', 'CUMAC001C','CUTE_001C','CURZA001C','CUVOS001C','0'])==True or all(x in list(df.iloc[:,i].unique()) for x in ['CUNIT001C', 'CUMAC001C','CURZA001C','CUVOS001C','0'])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'DS Base'})                        
                    elif all(x in list(df.iloc[:,i].unique()) for x in ['CUNIT001C','0'])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'DU'})                         
                    elif df.iloc[j,i][0:5]=='CECOL' or df.iloc[j,i][0:5]=='CELBA' or df.iloc[j,i][0:5]=='CEPRO':
                        df=df.rename(columns={lista_de_columna[i]:'DW FAAE_COMPOUND_MF'})    
                    elif df.iloc[j,i][0:3]=='SFA': 
                        df=df.rename(columns={lista_de_columna[i]:'FAAJ :ENCOLHIMENTO'})       
                        
                elif lista_de_columna[i]=='TEAB :ENCOLHIMENTO' or lista_de_columna[i]=='TEAB :ENCOLHIMENTO.1' :
                    if np.count_nonzero(np.array(df[lista_de_columna[i]].replace('0',0)))>770:
                        df=df.rename(columns={lista_de_columna[i]:'TELA 1 SAP'})
                    else:
                         df=df.rename(columns={lista_de_columna[i]:'TELA 2 SAP'})    
                            
                elif lista_de_columna[i]=='CNAA :ENCOLHIMENTO' or lista_de_columna[i]=='CNAA :ENCOLHIMENTO.1' :
                    if np.count_nonzero(np.array(df[lista_de_columna[i]].replace('0',0)))>770:
                        df=df.rename(columns={lista_de_columna[i]:'BELT SAP Belt2'})   
               
                if lista_de_columna[i]=='AVAA :ENCOLHIMENTO' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.1' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.2' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.3' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.4' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.5' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.6' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.7' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.8' or lista_de_columna[i]=='AVAA :ENCOLHIMENTO.9':                  
                    if (df.iloc[j,i][0:3]=='SAV' and len(df.iloc[j,i])>9): 
                        df=df.rename(columns={lista_de_columna[i]:'SAV code'})                       

            elif type(df.iloc[j,i])==float or type(df.iloc[j,i])==np.float64 or type(df.iloc[j,i])==np.int64 :#Verificar
                if lista_de_columna[i]=='AGAA :ENCOLHIMENTO' or lista_de_columna[i]=='AGAA :ENCOLHIMENTO.1' or lista_de_columna[i]=='AGAA :ENCOLHIMENTO.2':
                    if np.mean(np.array(df[lista_de_columna[i]].replace('0',0).unique()))>=3000:
                        df=df.rename(columns={lista_de_columna[i]:'Weight Carcass'})                      
                    
                if lista_de_columna[i]=='TEAB :ESPESOR MEDIO (CALCULADO' or lista_de_columna[i]=='TEAB :ESPESOR MEDIO (CALCULADO.1' :          
                    if all(x in list(df.iloc[:,i].unique()) for x in [0.9,1.05,1.2,1.3])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'AY TEAB :Ply thickness (calc)'})
                    if all(x in list(df.iloc[:,i].unique()) for x in [0.9,1.05])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'BC TEAB :Ply thickness (calc)'})            
                        
                #Como diferencia una tela de otra tela? 
                elif lista_de_columna[i]=='TEAB :ANGULO TELA' or lista_de_columna[i]=='TEAB :ANGULO TELA.1' or lista_de_columna[i]=='TEAB :ANGULO TELA.2' :
                    if all(x in list(df.iloc[:,i].unique()) for x in ['0',90,93])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'BD TEAB :Ply angle'})   
                    elif all(x in list(df.iloc[:,i].unique()) for x in [90,87])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'AZ TEAB :Ply angle'})                           

                                
                elif lista_de_columna[i]=='CNAA :ESPESOR MEDIO (CALCULADO' or lista_de_columna[i]=='CNAA :ESPESOR MEDIO (CALCULADO.1':          
                    if all(x in list(df.iloc[:,i].unique()) for x in [1.15,0.9])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'CNAA :Metallic belt thickness'})                 
            
                elif lista_de_columna[i]=='CNAA :ANGULO DE LA CINTURA' or lista_de_columna[i]=='CNAA :ANGULO DE LA CINTURA.1':                     
                    if all(x in list(df.iloc[:,i].unique()) for x in [25,27,30,32,35])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'CNAB :BELT ANGLE'})
                        
                elif lista_de_columna[i]=='CNAA :ANCHO DE LA CINTURA' or lista_de_columna[i]=='CNAA :ANCHO DE LA CINTURA.1':          
                    if all(x in list(df.iloc[:,i].unique()) for x in [155,160,165,170,175,180,185,190,195,200,205,210,215,220,225,230,235,240,245,250,255,260,265,270,275])==True:#Cambiar lista al actualizar tabla
                        df=df.rename(columns={lista_de_columna[i]:'CNAA :Metallic belt width'})  
                        
                if lista_de_columna[i]=='TEAB :ANCHO DE LA TELA' or lista_de_columna[i]=='TEAB :ANCHO DE LA TELA.1' :
                    if np.count_nonzero(np.array(df[lista_de_columna[i]].replace('0',0)))>770:#Revisar el tema de las columnas
                        df=df.rename(columns={lista_de_columna[i]:'AX TEAB :Ply width'})
                    else:
                         df=df.rename(columns={lista_de_columna[i]:'BB TEAB :Ply width'})    
                            
                    
            else:
                df=df.rename(columns={lista_de_columna[i]:lista_de_columna[i]})
            

    return df



#
def concatenar_columnas(df):
    #SAP
    df['FAAP :ENCOLHIMENTO']=df['FAAP :ENCOLHIMENTO'].replace('0','')
    df['FAAE :ENCOLHIMENTO']=df['FAAE :ENCOLHIMENTO'].replace('0','')
    df['FAAJ :ENCOLHIMENTO']=df['FAAJ :ENCOLHIMENTO'].replace('0','')
    df['CF Tread SAP code']=df['FAAP :ENCOLHIMENTO']+df['FAAE :ENCOLHIMENTO']+df['FAAJ :ENCOLHIMENTO']
    #DADOS
    df['CR Drawing code']=df['CR Drawing code'].replace('0','')
    df['DC Drawing code']=df['DC Drawing code'].replace('0','')
    df['DN Drawing code']=df['DN Drawing code'].replace('0','')
    df['CG Drawing code']=df['CR Drawing code']+df['DC Drawing code']+df['DN Drawing code']
    #ANCHO
    df['CS FAAE :Tread total width']=(df['CS FAAE :Tread total width'].replace('0','')).astype(str)
    df['DD FAAE :Tread total width']=(df['DD FAAE :Tread total width'].replace('0','')).astype(str)
    df['DO FAAE :Tread total width']=(df['DO FAAE :Tread total width'].replace('0','')).astype(str)
    df['CH FAAE :Tread total width']=df['CS FAAE :Tread total width']+df['DD FAAE :Tread total width']+df['DO FAAE :Tread total width']
    #ESPESOR BANDA DE RODAMIENTO
    df['CT FAAE :Tread shoulder thickness']=(df['CT FAAE :Tread shoulder thickness'].replace('0','')).astype(str)
    df['DE FAAE :Tread shoulder thickness']=(df['DE FAAE :Tread shoulder thickness'].replace('0','')).astype(str)
    df['DP FAAE :Tread shoulder thickness']=(df['DP FAAE :Tread shoulder thickness'].replace('0','')).astype(str)
    df['CI FAAE :Tread shoulder thickness']=df['CT FAAE :Tread shoulder thickness']+df['DE FAAE :Tread shoulder thickness']+df['DP FAAE :Tread shoulder thickness']  
    #ESPESOR 
    df['CU FAAP :Central thickness']=(df['CU FAAP :Central thickness'].replace('0','')).astype(str)
    df['DF FAAE :Tread centre thickness (']=(df['DF FAAE :Tread centre thickness ('].replace('0','')).astype(str)
    df['DQ FAAE :Tread centre thickness (']=(df['DQ FAAE :Tread centre thickness ('].replace('0','')).astype(str)
    df['CJ FAAP :Central thickness']=df['CU FAAP :Central thickness']+df['DF FAAE :Tread centre thickness (']+df['DQ FAAE :Tread centre thickness (']
    #COMPUESTO
    df['CV Cap Compound']=df['CV Cap Compound'].replace('0','')
    df['DG Cap Compound']=df['DG Cap Compound'].replace('0','')
    df['DR Cap Compound']=df['DR Cap Compound'].replace('0','')
    df['CK Cap Compound']=df['CV Cap Compound']+df['DG Cap Compound']+df['DR Cap Compound']
    #BASE
    df['CW Base']=df['CW Base'].replace('0','')
    df['DS Base']=df['DS Base'].replace('0','')
    df['CL Base']=df['CW Base']+df['DS Base']
    #BASE TKN
    df['FAAP :C_FAAP_UNDERLAY_THCK']=(df['FAAP :C_FAAP_UNDERLAY_THCK'].replace('0','')).astype(str)
    df['DT Foglietta']=(df['DT Foglietta'].replace('0','')).astype(str)
    df['CM Base TKN']=df['FAAP :C_FAAP_UNDERLAY_THCK']+df['DT Foglietta']
    #FOGLIETTA
    df['CY Foglietta']=df['CY Foglietta'].replace('0','')
    df['DJ Foglietta']=df['DJ Foglietta'].replace('0','')
    df['DU']=df['DU'].replace('0','')
    df['CN Foglietta']=df['CY Foglietta']+df['DJ Foglietta']+df['DU']
    #FOGLIETTA TKN
    df['FAAP :C_FAAP_UNDERC_THCK']=(df['FAAP :C_FAAP_UNDERC_THCK'].replace('0','')).astype(str)
    df['FAAE :ESPESSURA DA FOLHETA (SU']=(df['FAAE :ESPESSURA DA FOLHETA (SU'].replace('0','')).astype(str)
    df['DV']=(df['DV'].replace('0','')).astype(str)
    df['Foglietta TKN']=df['FAAP :C_FAAP_UNDERC_THCK']+df['FAAE :ESPESSURA DA FOLHETA (SU']+df['DV']   
    #COMPUESTO
    df['DA FAAP_COMPOUND_MF']=df['DA FAAP_COMPOUND_MF'].replace('0','')
    df['DL FAAE_COMPOUND_MF']=df['DL FAAE_COMPOUND_MF'].replace('0','')
    df['DW FAAE_COMPOUND_MF']=df['DW FAAE_COMPOUND_MF'].replace('0','')
    df['CP FAAP_COMPOUND_MF']=df['DA FAAP_COMPOUND_MF']+df['DL FAAE_COMPOUND_MF']+df['DW FAAE_COMPOUND_MF']
    
    
    #AE Complex SAP
    df['Complex SAP']=df['Complex SAP'].replace('0','')
    df['CPAP :ENCOLHIMENTO']=df['CPAP :ENCOLHIMENTO'].replace('0','')
    df['CPAC :ENCOLHIMENTO']=df['CPAC :ENCOLHIMENTO'].replace('0','')
    df['Complex SAP']=df['Complex SAP']+df['CPAP :ENCOLHIMENTO']+df['CPAC :ENCOLHIMENTO']
    #AF Total width
    df['Total width']=df['Total width'].replace('0','')
    df['CPAP :C_CPAP_TOT_WDTH']=df['CPAP :C_CPAP_TOT_WDTH'].replace('0','')
    df['CPAC :ANCHO DEL RELLENO']=df['CPAC :ANCHO DEL RELLENO'].replace('0','')
    df['Total width']=df['Total width'].astype(str)+df['CPAP :C_CPAP_TOT_WDTH'].astype(str)+df['CPAC :ANCHO DEL RELLENO'].astype(str)
    #AG CPAA_COMPOUND_CH
    df['CPAA_COMPOUND_CH']= df['CPAA_COMPOUND_CH'].replace('0','')
    df['CPAP :C_CPAP_STRIPC_CMPD_K']=df['CPAP :C_CPAP_STRIPC_CMPD_K'].replace('0','')
    df['CPAC :CODIGO DE MEZCLA DE RELL']=df['CPAC :CODIGO DE MEZCLA DE RELL'].replace('0','')
    df['CPAA_COMPOUND_CH']=df['CPAA_COMPOUND_CH'].astype(str)+df['CPAP :C_CPAP_STRIPC_CMPD_K'].astype(str)+df['CPAC :CODIGO DE MEZCLA DE RELL'].astype(str)
    #AH CPAA :Complex cushion thicknes
    df['CPAA :Complex cushion thicknes']=df['CPAA :Complex cushion thicknes'].replace('0','')
    df['CPAP :C_CPAP_STRIPC_THCK']=df['CPAP :C_CPAP_STRIPC_THCK'].replace('0','')
    df['CPAC :ESPESOR DEL RELLENO']=df['CPAC :ESPESOR DEL RELLENO'].replace('0','')
    df['CPAA :Complex cushion thicknes']=df['CPAA :Complex cushion thicknes'].astype(str)+df['CPAP :C_CPAP_STRIPC_THCK'].astype(str)+df['CPAC :ESPESOR DEL RELLENO'].astype(str)
    #AI CPAA :Complex cushion width
    df['CPAA :Complex cushion width']=df['CPAA :Complex cushion width'].replace('0','')
    df['CPAP :C_CPAP_LIN_WDTH']=df['CPAP :C_CPAP_LIN_WDTH'].replace('0','')
    df['CPAC :ANCHO DE LINER']=df['CPAC :ANCHO DE LINER'].replace('0','')
    df['CPAA :Complex cushion width']=df['CPAA :Complex cushion width'].astype(str)+df['CPAP :C_CPAP_LIN_WDTH'].astype(str)+df['CPAC :ANCHO DE LINER'].astype(str)
    #AJ CPAA_COMPOUND_LN
    df['CPAA_COMPOUND_LN']=df['CPAA_COMPOUND_LN'].replace('0','')
    df['CPAP :C_CPAP_LIN_CMPD_K']=df['CPAP :C_CPAP_LIN_CMPD_K'].replace('0','')
    df['CPAC :MEZCLA DE LINER']=df['CPAC :MEZCLA DE LINER'].replace('0','')
    df['CPAA_COMPOUND_LN']=df['CPAA_COMPOUND_LN'].astype(str)+df['CPAP :C_CPAP_LIN_CMPD_K'].astype(str)+df['CPAC :MEZCLA DE LINER'].astype(str)
    #AK CPAA :Complex liner thickness
    df['CPAA :Complex liner thickness']=df['CPAA :Complex liner thickness'].replace('0','')
    df['CPAP :C_CPAP_LIN_THCK']=df['CPAP :C_CPAP_LIN_THCK'].replace('0','')
    df['CPAC :ESPESOR DE LINER']=df['CPAC :ESPESOR DE LINER'].replace('0','')
    df['CPAA :Complex liner thickness']=df['CPAA :Complex liner thickness'].astype(str)+df['CPAP :C_CPAP_LIN_THCK'].astype(str)+df['CPAC :ESPESOR DE LINER'].astype(str)
    #AL CPAA :Complex liner width
    df['CPAA :Complex liner width']=df['CPAA :Complex liner width'].replace('0','')
    df['CPAP :C_CPAP_STRIPC_WDTH']=df['CPAP :C_CPAP_STRIPC_WDTH'].replace('0','')
    df['CPAA :Complex liner width']=df['CPAA :Complex liner width'].astype(str)+df['CPAP :C_CPAP_STRIPC_WDTH'].astype(str)
    #AM CPAA_COMPOUND_LT
    df['CPAA_COMPOUND_LT']=df['CPAA_COMPOUND_LT'].replace('0','')
    df['CPAP :C_CPAP_HEEL_CMPD_K']= df['CPAP :C_CPAP_HEEL_CMPD_K'].replace('0','')
    df['CPAA_COMPOUND_LT']=df['CPAA_COMPOUND_LT'].astype(str)+df['CPAP :C_CPAP_HEEL_CMPD_K'].astype(str)
    #AN CPAA :Bead support strip thick
    df['CPAA :Bead support strip thick']=df['CPAA :Bead support strip thick'].replace('0','')
    df['CPAP :C_CPAP_HEEL_THCK']=df['CPAP :C_CPAP_HEEL_THCK'].replace('0','')
    df['CPAA :Bead support strip thick']=df['CPAA :Bead support strip thick'].astype(str)+df['CPAP :C_CPAP_HEEL_THCK'].astype(str)
    #AO CPAA :Bead support strip width
    df['CPAA :Bead support strip width']=df['CPAA :Bead support strip width'].replace('0','')
    df['CPAP :C_CPAP_HEEL_WDTH']=df['CPAP :C_CPAP_HEEL_WDTH'].replace('0','')
    df['CPAA :Bead support strip width']=df['CPAA :Bead support strip width'].astype(str)+df['CPAP :C_CPAP_HEEL_WDTH'].astype(str)

        
    return df

def asignar_datos(a):
    z=''
    if a=='0':
        z=''
    else:
        z='YES'
    return z
    
def ty_std(a):
    z=''
    if a=='0':
        z='STD'
    else:
        z='RUN-FLAT'
    return z

def p_column(a):
    z=''
    if a[0:1]=='P':
        z='P'
    else:
        z=''
    return z

def flipper(a):
    z=''
    if a==45:
        z='Flipper'
    else:
        z=''
    return z


def nombrar_columnas(df):
    
    df=df.rename(columns={'AVAA :CÓDIGO IP ( 5 POSICIONES':'IP code','AGAA :Tread name':'Pattern',
        'AVAA :Product Specification nu':'AVAA :Product Specification nu','AVAA :TUNING':'AVAA :TUNING',
        'AGAA :ESQUEMA TELAS (1+0, 1+1,':'Ply sch','AGAA :A_SO_AGAA_DIM_WIDTH':'width','AGAA :A_SO_AGAA_DIM_RATIO':'ratio',
        'AGAA :RIM_(CALENTAMIENTO)':'rim','AGAA :Load Index':'LI','AGAA :A_SO_AGAA_SPEED_INDEX_1':'SI','AVAA :A_SO_AVAA_ORIGINAL_EQUIP':
        'CUSTOMER','AVAA :':'WEIGHT','ACAA :A_SO_A70F_DRUM_WIDTH':'Drum Wdth','AGAA :DIÁMETRO DEPOSITO 1RACIN':
        'Aux drum diameter','AGAA :A_SO_AGAA_M_AG_CIRCUMF':'Green tyre circumf.','AGAA :EPI Sectors':'EPI Sectors','AGAA :EPI Spacer':
        'EPI Spacer','ACAA :ESQUEMA REF.TALON/ANT':'Chafer scheme','BRAB :ANGULO':'Chafer angle','BRAB :LARGURA DO TECIDO':'Chafer width','CPAA :C_CPAA_TOT_WDTH':'Total width','CPAA :CODIGO DE MEZCLA DE RELL':'CPAA_COMPOUND_CH','CPAA :ESPESOR DEL RELLENO':
        'CPAA :Complex cushion thicknes','CPAA :ANCHO DEL RELLENO':'CPAA :Complex cushion width','CPAA :MEZCLA DE LINER':'CPAA_COMPOUND_LN',
        'CPAA :ESPESOR DE LINER':'CPAA :Complex liner thickness','CPAA :ANCHO DE LINER':'CPAA :Complex liner width','CPAA :C_CPAA_STRIPE_CMPD_K':
        'CPAA_COMPOUND_LT','CPAA :ESPESOR LISTA BASE TALON':'CPAA :Bead support strip thick','CPAA :ANCHO LISTA BASE TALON':
        'CPAA :Bead support strip width','CPAP :A_CPAP_A_DISTANCE':'GAP','LUAA :A_SO_LUAA_DIE_CODE':'Lunetta die','LUAA :Espesor listino':
        'Lunetta Thck','LUAA :Ancho listino':'Lunetta Wdth','ACAA :A_SO_ACAA_D_BUILD_UP_SCH':'Dist. Between Lunette','CEBM :A_SO_CEBM_DRAWING':'CEBM :Drawing','CEBM :A_SO_CEBM_PROFILE':
        'CEBM :Filler profile','CEBM :C_CEBM_BASE':'CEBM :Filler base width','CEBM :C_CEBM_HEIGHT':'CEBM :C_CEBM_HEIGHT','CEBH :ENCOLHIMENTO':
        'CEJA SAP Aro','CEBH :C_CEBH_DEVPT':'CEBM :Bead development','CEBM :C_CEBM_NOM_RIM':'CEBM :Bead nominal rim diamete',
        'ACAA :A_SO_A70F_BEAD_SET_RING_':'ACAA :Bead se Ring Diameter','CEBM :C_CEBM_TURNS_SEQUENCE':'Scheme','FLAA :ANGULO CUBRE-TALON':'Flipp °','FLAA :ANCHO CUBRE-TALON':'Flipp Width','BSAA :A_SO_BSAA_DESCRIPT':'BSAA :Description','BSAA :ESPESSURA MÉDIA (CALCULA':'Bandina Thick','AGAA :A_SO_AGAA_BANDINA_TOT_WD':
        'AGAA :Bandina total width','AGAA :A_SO_AGAA_D_DRAWING':'AGAA :Bandina Drawing','AGAA :A_SO_AGAA_D_WINDING_SCHE':'AGAA :Winding scheme',
        'CNAB :ENCOLHIMENTO':'BELT1 SAP Belt+listini','CNAB :A_SO_CNAB_DESCRIPT':'CNAB :Description','CNAB :ANCHO DE LA CINTURA':'CNAB :BELT WIDTH',
        'FAAP :C_FAAP_DIE_CODE_MAIN':'CR Drawing code','FAAP :Ancho total':'CS FAAE :Tread total width','FAAP :A_SO_FAAP_FA_SHOU_THCK':
        'CT FAAE :Tread shoulder thickness','FAAP :Espesor centro':'CU FAAP :Central thickness','FAAE :C_FAAE_DIE_CODE_MAIN':'DC Drawing code',
        'FAAE :LARGURA TOTAL':'DD FAAE :Tread total width','FAAE :ESPESSURA LATERAL':'DE FAAE :Tread shoulder thickness','FAAE :ESPESSURA MÉDIA (CALCULA':
        'DF FAAE :Tread centre thickness (','FAAJ :A_FAAJ_DIE_CODE':'DN Drawing code','FAAJ :C_FAAJ_TOT_WDTH':'DO FAAE :Tread total width',
        'FAAJ :C_FAAJ_TOT_THCK':'DP FAAE :Tread shoulder thickness','FAAJ :Espesor centro':'DQ FAAE :Tread centre thickness (',
        'FAAJ :ESP 1º FOLHETA SOB RODAG':'DT Foglietta','FAAJ :ESP 2º FOLHETA SOB RODAG':'DV','FNAB :C_FNAB_FN_DIE_CODE_MAIN':
        'FNAB :Drawing code','FNAB :ANCHO TOTAL':'Total sidewall assembly','FNAB :ANCHO DE LO COSTADO':'SW Width','FNAB :ESPESOR DE LO COSTADO':
        'SW Thick','FNAB :ANCHO LISTA ANTI-ABRASIV':'Heelstrip Width','FNAB :ESPESOR LISTA ANTI-ABRAS':'Heelstrip Thick','AGAA :A_SO_AGAA_SIDEW_DIST_DRU':
        'AGAA :A_SO_AGAA_SIDEW_DIST_DRU','AVAA :A_SO_HF52_BLADDER_SIZE':'Bladder size','AVAA :Container code':'Container code',
        'AVAA :A_SO_HF52_CURING_SCHEME_':'Curing Scheme','AVAA :A_SO_HF52_CURING_TIME':'Curing time','AVAA :A_SO_HF52_CUR_SCHEME_RDC':
        'Curing Scheme RDC','AVAA :A_SO_HF52_CUR_TIME_RDC':'Curing Time RDC','AVAA :A_SO_HF52_PCI':'PCI On/Off','AVAA :A_SO_HF52_PCI_TYPE':
        'PCI Type','AVAA :A_SO_HF52_PCI_WDTH':'PCI Width','AVAA :A_SO_HF52_PRES_PCI':'PCI Pressure','AVAA :A_SO_HF52_PRES_CONT':'AVAA :Presure container [bar]',
        'AVAA :A_SO_HF52_PRES_PLATE':'AVAA :Presure plate [bar]','AVAA :A_SO_HF52_PRES_STEAM':'AVAA :Pressure steam [bar]','AVAA :A_SO_HF52_SECTOR_DRW':
        'Sector drawing','AVAA :A_SO_HF52_SIDEWALL_DRW':'Sidewall drawing','AVAA :A_SO_HF52_TEMP_CONT':'AVAA :Temperature container [°C]',
        'AVAA :A_SO_HF52_TEMP_PLATE':'AVAA :Temperature plate [°C]','AVAA :A_SO_AVAA_I_RECALL_NR':'DOT'})

            
    df_1=concatenar_columnas(df)
    df_1['AP']=df_1['GAP'].apply(asignar_datos)
    df_1['Type']=df_1['Lunetta die'].apply(ty_std)
    df_1['L']=df_1['MISURA'].apply(p_column)
    df_1['Flipper']=df_1['Flipp °'].apply(flipper)
    
   
    modificaciones=df_1[['IP code','MISURA','Pattern','AVAA :Product Specification nu','AVAA :TUNING','AVAA :Tyre INDST',
        'Ply sch','Type','L','width','ratio','rim','LI','SI','CUSTOMER','WEIGHT','Weight Carcass','Drum Wdth','Aux drum diameter','Green tyre circumf.',  
        'EPI Sectors','EPI Spacer','Chafer scheme','Chafer angle','Chafer width','Complex SAP','Total width','CPAA_COMPOUND_CH','CPAA :Complex cushion thicknes','CPAA :Complex cushion width',
        'CPAA_COMPOUND_LN','CPAA :Complex liner thickness','CPAA :Complex liner width','CPAA_COMPOUND_LT','CPAA :Bead support strip thick',
        'CPAA :Bead support strip width','AP','GAP','Lunetta die','Lunetta Thck','Lunetta Wdth','Dist. Between Lunette','TELA 1 SAP','TEAB :Description','AX TEAB :Ply width','AY TEAB :Ply thickness (calc)',
        'AZ TEAB :Ply angle','TELA 2 SAP','BB TEAB :Ply width','BC TEAB :Ply thickness (calc)','BD TEAB :Ply angle','CEJA SAP Aro+filler','CEBM :Drawing','CEBM :Filler profile','CEBM_COMPOUND_RC',
        'CEBM :Filler base width','CEBM :C_CEBM_HEIGHT','CEJA SAP Aro','CEBM :Bead development','CEBM :Bead nominal rim diamete','ACAA :Bead se Ring Diameter',
        'Scheme','Flipper','Flipp °','Flipp Width','BSAA :Description','Bandina Thick','AGAA :Bandina total width','AGAA :Bandina Drawing','AGAA :Winding scheme',
        'BELT1 SAP Belt+listini','CNAB :Description','CNAB :BELT WIDTH','BELT SAP Belt2','CNAA :Metallic belt width','CNAB :BELT ANGLE'
        ,'CNAA :Metallic belt thickness','CF Tread SAP code','CG Drawing code','CH FAAE :Tread total width','CI FAAE :Tread shoulder thickness','CJ FAAP :Central thickness','CK Cap Compound','CL Base','CM Base TKN','CN Foglietta','Foglietta TKN','CP FAAP_COMPOUND_MF','FAAP :ENCOLHIMENTO',                                 
        'CR Drawing code','CS FAAE :Tread total width','CT FAAE :Tread shoulder thickness','CU FAAP :Central thickness','CV Cap Compound','CW Base',
        'FAAP :C_FAAP_UNDERLAY_THCK','CY Foglietta','FAAP :C_FAAP_UNDERC_THCK','DA FAAP_COMPOUND_MF','FAAE :ENCOLHIMENTO',
        'DC Drawing code','DD FAAE :Tread total width','DE FAAE :Tread shoulder thickness','DF FAAE :Tread centre thickness (','DG Cap Compound',
        'DJ Foglietta','FAAE :ESPESSURA DA FOLHETA (SU','DL FAAE_COMPOUND_MF','FAAJ :ENCOLHIMENTO','DN Drawing code','DO FAAE :Tread total width','DP FAAE :Tread shoulder thickness',
        'DQ FAAE :Tread centre thickness (','DR Cap Compound','DS Base','DT Foglietta','DU','DV','DW FAAE_COMPOUND_MF','FNAB :ENCOLHIMENTO',
        'FNAB :Drawing code','Total sidewall assembly','SW Width','SW Thick','Heelstrip Width','Heelstrip Thick','FNAB_COMPOUND_FN','FNAB_COMPOUND_AB',
        'AGAA :A_SO_AGAA_SIDEW_DIST_DRU','Bladder size','Container code','Curing Scheme','Curing time','Curing Scheme RDC','Curing Time RDC',
        'PCI On/Off','PCI Type','PCI Width','PCI Pressure','AVAA :Presure container [bar]','AVAA :Presure plate [bar]','AVAA :Pressure steam [bar]',
        'Sector drawing','Sidewall drawing','AVAA :Temperature container [°C]','AVAA :Temperature plate [°C]','DOT','AVAA :Sponge Code (L1)','SAV code'
        ]]
    return modificaciones

def pintar_celda(wb):
    ws=wb['hoja1']
       
    #color VERDE en celdas que contienen nombres columnas; rango B-DB
    fill_pattern = PatternFill(patternType='solid',fgColor='00FF00') 
    
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=106):
        for cell in row:
            cell.fill = fill_pattern
    
    #color AZUL en celdas que contienen nombres de columnas; rango DC-DX        
    fill_pattern = PatternFill(patternType='solid',fgColor='00B0F0')         
    
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=107, max_col=128):
        for cell in row:
            cell.fill = fill_pattern  
    
    #color AMARILLO en celdas que contienen nombres de columnas; rango I1        
    fill_pattern = PatternFill(patternType='solid',fgColor='CCFF33')
    ws['I1'].fill=fill_pattern
    
                  
    #color LILA ; rango I2-I491       
    fill_pattern = PatternFill(patternType='solid',fgColor='F2DCDB')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=9, max_col=9):
        for cell in row:
            cell.fill = fill_pattern  
            
    #color GRIS AZULADO ;rango X2-AB491      
    fill_pattern = PatternFill(patternType='solid',fgColor='BFBFBF')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=24, max_col=28):
        for cell in row:
            cell.fill = fill_pattern  
            
    #color ANARANJADO-claro ;rango AF2-AR491     
    fill_pattern = PatternFill(patternType='solid',fgColor='FCD5B4')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=32, max_col=44):
        for cell in row:
            cell.fill = fill_pattern 
    
    #color ANARANJADO ;rango AG2-AG491      
    fill_pattern = PatternFill(patternType='solid',fgColor='FFCC99')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=33, max_col=33):
        for cell in row:
            cell.fill = fill_pattern  
            
    #color ANARANJADO/CLARO ;rango AW2-BE491       
    fill_pattern = PatternFill(patternType='solid',fgColor='FCD5B4')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=49, max_col=57):
        for cell in row:
            cell.fill = fill_pattern 
            
    #color ANARANJADO/CLARO ;rango BT2-BX491      
    fill_pattern = PatternFill(patternType='solid',fgColor='FCD5B4')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=72, max_col=76):
        for cell in row:
            cell.fill = fill_pattern      
            
    #color GRIS; rango CF2-CF491      
    fill_pattern = PatternFill(patternType='solid',fgColor='D8E4BC')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=84, max_col=84):
        for cell in row:
            cell.fill = fill_pattern    
            
    #color ANARANJADO/CLARO; rango CG2-CQ491       
    fill_pattern = PatternFill(patternType='solid',fgColor='FCD5B4')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=85, max_col=95):
        for cell in row:
            cell.fill = fill_pattern 
            
            
    #color GRIS; rango CR2-CS491      
    fill_pattern = PatternFill(patternType='solid',fgColor='D8E4BC')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=96, max_col=97):
        for cell in row:
            cell.fill = fill_pattern         
                    
    #color GRIS AZULADO; rango DC2-DV491       
    fill_pattern = PatternFill(patternType='solid',fgColor='BFBFBF')         
    
    for row in ws.iter_rows(min_row=2, max_row=491, min_col=107, max_col=126):
        for cell in row:
            cell.fill = fill_pattern  
            
    #COLOR DE LETRAS
    #color AZUL negritas; rango B2-B700  
    font_style = Font(bold=True,color="0000FF")
    for row in ws.iter_rows(min_row=2, max_row=700, min_col=2, max_col=2):
        for cell in row:
            cell.font = font_style  
            
    #color ROJO; rango C2-C700  
    font_style = Font(color="FF0000")
    for row in ws.iter_rows(min_row=2, max_row=700, min_col=3, max_col=3):
        for cell in row:
            cell.font = font_style       
            
    #color AZUL; rango C2-C700 
    font_style = Font(color="0000FF")
    for row in ws.iter_rows(min_row=2, max_row=700, min_col=4, max_col=157):
        for cell in row:
            cell.font = font_style 

    #ALINEAR TEXTO 
    rows = range(1, 700)
    columns = range(1, 128)
    for row in rows:
        for col in columns:
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')# wrap_text=True)               

    return wb


def orden_final(df_1):    
    modificaciones=df_1[['IP code','Resp','MISURA','Pattern','AVAA :Product Specification nu','AVAA :TUNING','AVAA :Tyre INDST','SumOfProd_MP2021',
        'Ply sch','SW scheme','Type','L','width','ratio','rim','LI','SI','CUSTOMER','REPL / OE','OE Program','WEIGHT','Weight Carcass','Drum Wdth','Aux drum diameter','Green tyre circumf.',  
        'EPI Sectors','EPI Spacer','Chafer scheme','Chafer angle','Chafer width','Complex SAP','Total width','CPAA_COMPOUND_CH','CPAA :Complex cushion thicknes','CPAA :Complex cushion width',
        'CPAA_COMPOUND_LN','CPAA :Complex liner thickness','CPAA :Complex liner width','CPAA_COMPOUND_LT','CPAA :Bead support strip thick',
        'CPAA :Bead support strip width','AP','GAP','Lunetta die','Lunetta Thck','Lunetta Wdth','Dist. Between Lunette','TELA 1 SAP','TEAB :Description','AX TEAB :Ply width'
        ,'AY TEAB :Ply thickness (calc)',
        'AZ TEAB :Ply angle',
        'TELA 2 SAP','BB TEAB :Ply width',
        'BC TEAB :Ply thickness (calc)','BD TEAB :Ply angle','CEJA SAP Aro+filler','CEBM :Drawing','CEBM :Filler profile','CEBM_COMPOUND_RC',
        'CEBM :Filler base width','CEBM :C_CEBM_HEIGHT','CEJA SAP Aro','CEBM :Bead development','CEBM :Bead nominal rim diamete','ACAA :Bead se Ring Diameter',
        'Scheme','Flipper','Flipp °','Flipp Width','BSAA :Description','Bandina Thick','AGAA :Bandina total width','AGAA :Bandina Drawing','AGAA :Winding scheme',
        'BELT1 SAP Belt+listini','CNAB :Description','CNAB :BELT WIDTH','BELT SAP Belt2','CNAA :Metallic belt width','CNAB :BELT ANGLE'
        ,'CNAA :Metallic belt thickness','Tread Area',                                
        'CF Tread SAP code','CG Drawing code','CH FAAE :Tread total width','CI FAAE :Tread shoulder thickness','CJ FAAP :Central thickness','CK Cap Compound','CL Base','CM Base TKN','CN Foglietta','Foglietta TKN','CP FAAP_COMPOUND_MF',
        'SW Area','FNAB :ENCOLHIMENTO',
        'FNAB :Drawing code','Total sidewall assembly','SW Width','SW Thick','Heelstrip Width','Heelstrip Thick','FNAB_COMPOUND_FN','FNAB_COMPOUND_AB',
        'AGAA :A_SO_AGAA_SIDEW_DIST_DRU','Bladder size','Container code','Curing Scheme','Curing time','Curing Scheme RDC','Curing Time RDC',
        'PCI On/Off','PCI Type','PCI Width','PCI Pressure','AVAA :Presure container [bar]','AVAA :Presure plate [bar]','AVAA :Pressure steam [bar]',
        'Sector drawing','Sidewall drawing','AVAA :Temperature container [°C]','AVAA :Temperature plate [°C]','DOT','AVAA :Sponge Code (L1)','AVAA :Sponge Description','SAV code'
        ]]
    return modificaciones
    

def abreviacion_compuestos(tabla_modificada):
    #ABREVIACIÓN DE COMPUESTOS 
    #CPAA_COMPOUND_CH
    tabla_modificada['CPAA_COMPOUND_CH']=(tabla_modificada['CPAA_COMPOUND_CH'].replace ('CLMAU001C','LMAU')).astype(str) 
    tabla_modificada['CPAA_COMPOUND_CH']=(tabla_modificada['CPAA_COMPOUND_CH'].replace ('CLAUF001C','LAUF')).astype(str)
    #CPAA_COMPOUND_LN
    tabla_modificada['CPAA_COMPOUND_LN']=(tabla_modificada['CPAA_COMPOUND_LN'].replace ('CLAME001C','LAME')).astype(str) 
    #CPAA_COMPOUND_LT
    tabla_modificada['CPAA_COMPOUND_LT']=(tabla_modificada['CPAA_COMPOUND_LT'].replace ('CTOM_001C','TOM')).astype(str) 
    tabla_modificada['CPAA_COMPOUND_LT']=(tabla_modificada['CPAA_COMPOUND_LT'].replace ('CTURN001C','TURN')).astype(str) 
    tabla_modificada['CPAA_COMPOUND_LT']=(tabla_modificada['CPAA_COMPOUND_LT'].replace ('CTLC_001C','TLC')).astype(str) 
    #CEBM_COMPOUND_RC
    tabla_modificada['CEBM_COMPOUND_RC']=(tabla_modificada['CEBM_COMPOUND_RC'].replace ('CTAMI001C','TAMI')).astype(str) 
    tabla_modificada['CEBM_COMPOUND_RC']=(tabla_modificada['CEBM_COMPOUND_RC'].replace ('CTILO001C','TILO')).astype(str) 
    tabla_modificada['CEBM_COMPOUND_RC']=(tabla_modificada['CEBM_COMPOUND_RC'].replace ('CTHOR001C','THOR')).astype(str) 
    #CK Cap Compound
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDLUM001C','DLUM')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CBANK001C','BANK')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDRIE001C','DRIE')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CBERO001C','BERO')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDCAM001C','DCAM')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDTCM001C','DTCM')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDASR001C','DASR')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDWEM001C','DWEM')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CBODA001C','BODA')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDFLI001C','DFLI')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDOME001C','DOME')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDUFF001C','DUFF')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDJPB001C','DJPB')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDFUM001C','DFUM')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CBIEN001C','BIEN')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CBOOX001C','BOOX')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('C1VV896AC','1VV8')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDAJE001C','DAJE')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDUMI001C','DUMI')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDIEG001C','DIEG')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('CDFOG001C','DFOG')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('C1RW374WC','1RW3')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('C1VZ932EC','1VZ9')).astype(str) 
    tabla_modificada['CK Cap Compound']=(tabla_modificada['CK Cap Compound'].replace ('C1VW530GC','1VW5')).astype(str) 
    #CL Base 
    tabla_modificada['CL Base']=(tabla_modificada['CL Base'].replace ('CUTE_001C','UTE')).astype(str) 
    tabla_modificada['CL Base']=(tabla_modificada['CL Base'].replace ('CUNIT001C','UNIT')).astype(str) 
    tabla_modificada['CL Base']=(tabla_modificada['CL Base'].replace ('CUMAC001C','UMAC')).astype(str) 
    tabla_modificada['CL Base']=(tabla_modificada['CL Base'].replace ('CUVOS001C','UVOS')).astype(str) 
    tabla_modificada['CL Base']=(tabla_modificada['CL Base'].replace ('CURZA001C','URZA')).astype(str) 
    #CN Foglietta
    tabla_modificada['CN Foglietta']=(tabla_modificada['CN Foglietta'].replace ('CUNIT001C','UNIT')).astype(str) 
    tabla_modificada['CN Foglietta']=(tabla_modificada['CN Foglietta'].replace ('CUMAC001C','UMAC')).astype(str) 
    #CP FAAP_COMPOUND_MF
    tabla_modificada['CP FAAP_COMPOUND_MF']=(tabla_modificada['CP FAAP_COMPOUND_MF'].replace ('CELBA001C','ELBA')).astype(str) 
    tabla_modificada['CP FAAP_COMPOUND_MF']=(tabla_modificada['CP FAAP_COMPOUND_MF'].replace ('CECOL001C','ECOL')).astype(str) 
    #FNAB_COMPOUND_FN
    tabla_modificada['FNAB_COMPOUND_FN']=(tabla_modificada['FNAB_COMPOUND_FN'].replace ('CEPRO001C','EPRO')).astype(str) 
    tabla_modificada['FNAB_COMPOUND_FN']=(tabla_modificada['FNAB_COMPOUND_FN'].replace ('CELBA001C','ELBA')).astype(str) 
    tabla_modificada['FNAB_COMPOUND_FN']=(tabla_modificada['FNAB_COMPOUND_FN'].replace ('CECOL001C','ECOL')).astype(str) 
    #FNAB_COMPOUND_AB
    tabla_modificada['FNAB_COMPOUND_AB']=(tabla_modificada['FNAB_COMPOUND_AB'].replace ('CTURN001C','TURN')).astype(str) 
    tabla_modificada['FNAB_COMPOUND_AB']=(tabla_modificada['FNAB_COMPOUND_AB'].replace ('CTLC_001C','TLC')).astype(str) 
    return tabla_modificada