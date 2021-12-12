# -*- coding: utf-8 -*-
"""
Created on Fri Oct  8 14:11:10 2021

@author: Fernando Isai Almaraz Fabián 
"""
#Cargar librerias necesarias
import tkinter as tk    
from functools import partial
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl import load_workbook
import funciones_totales_copy2 as ft


#Generación de las funciones a utilizar
def getvalue(mystring,lista_1):
    a=mystring.get()
    lista_1.append(a)
    print(lista_1)
    


def funcion_principal(liga1_mod,liga2_mod,liga3_mod):
       
    try:
        #Importación de tablas necesarias para trabajar
        tabla_base= pd.read_excel(liga1_mod[0],header=0)
        areabdr= pd.read_excel(liga3_mod[0],sheet_name='AREABDR',header=0)
        areasw= pd.read_excel(liga3_mod[0],sheet_name='AREASW',header=0)
        oe= pd.read_excel(liga3_mod[0],sheet_name='OE',header=0)
        esponja= pd.read_excel(liga3_mod[0],sheet_name='ESPONJA',header=0)
        remplazo=pd.read_excel(liga3_mod[0],sheet_name='REMPLAZO',header=0)
        R=pd.read_excel(liga3_mod[0],sheet_name='R',header=0)
        TOS=pd.read_excel(liga3_mod[0],sheet_name='TOS',header=0)
        RESP=pd.read_excel(liga3_mod[0],sheet_name='RESP',header=0)
        
        #Limpiando los datos
        tabla_base=tabla_base.replace(np.nan,'0')
        
        #Renombrando las columnas
        tabla_modificada=ft.renombrando_columna(tabla_base)
        tabla_modificada=tabla_modificada.loc[:,~tabla_modificada.columns.duplicated()]
        modificaciones=ft.nombrar_columnas(tabla_modificada)
        modificaciones_sin_duplicados=modificaciones.drop_duplicates(subset='IP code')
        modificaciones_compuestos=ft.abreviacion_compuestos(modificaciones_sin_duplicados)
        
        
        #MUniendo las tablas
        modificaciones_compuestos_bdr=pd.merge(modificaciones_compuestos,areabdr,on='CG Drawing code',how='left')
        modificaciones_compuestos_sw=pd.merge(modificaciones_compuestos_bdr,areasw,on='FNAB :Drawing code',how='left')
        modificaciones_compuestos_oe=pd.merge(modificaciones_compuestos_sw,oe,on='IP code',how='left')
        modificaciones_compuestos_esponja=pd.merge(modificaciones_compuestos_oe,esponja,on='IP code',how='left')
        modificaciones_compuestos_remplazo=pd.merge(modificaciones_compuestos_esponja,remplazo,on='IP code',how='left')
        modificaciones_compuestos_R=pd.merge(modificaciones_compuestos_remplazo,R,on='IP code',how='left')
        modificaciones_compuestos_TOS=pd.merge(modificaciones_compuestos_R,TOS,on='IP code',how='left')
        modificaciones_compuestos_RESP=pd.merge(modificaciones_compuestos_TOS,RESP,on='IP code',how='left')

        modificaciones_compuestos_R_2=modificaciones_compuestos_RESP.dropna(subset=['SumOfProd_MP2021'])
        modificaciones_compuestos_R_3=ft.orden_final(modificaciones_compuestos_R_2)
        modificaciones_compuestos_R_3=modificaciones_compuestos_R_3.drop_duplicates(subset='IP code')
        modificaciones_compuestos_R_3=modificaciones_compuestos_R_3.sort_values(by='IP code')

        
        writer=pd.ExcelWriter(liga2_mod[0])
        modificaciones_compuestos_R_3.to_excel(writer,sheet_name='hoja1')
        writer.save()
        writer.close()
        wb = load_workbook(liga2_mod[0])
        tabla_pintada=ft.pintar_celda(wb)
        wb.save(liga2_mod[0])
        
    except IndexError:
        print('Aun no existe ningun elemento en la lista')
        

#Definición de la ventana principal
root =tk.Tk()
root.title('Descomplexity')
root.configure(background='#5A5656')
#validación de variables de entrada
mystring =tk.StringVar(root)
mystring2=tk.StringVar(root)
mystring3=tk.StringVar(root)

#creación de listas
lista_1=[]
lista_2=[]
lista_3=[]

#Generación de caja de entrada.
e1 = tk.Entry(root,textvariable = mystring,width=100,fg="blue",bd=3,selectbackground='Violet').pack()


#Generación de boton
button1 = tk.Button(root, 
                text='LIGA DE ENTRADA', 
                fg='White', 
                bg= 'dark green',height = 1, width = 20,command=partial(getvalue,mystring,lista_1)).pack()

#Generación de caja de entrada.
e2 = tk.Entry(root,textvariable = mystring2,width=100,fg="blue",bd=3,selectbackground='Violet').pack()

#Generación de boton
button2 = tk.Button(root, 
                text='LIGA DE SALIDA', 
                fg='White', 
                bg= 'dark green',height = 1, width = 20,command=partial(getvalue,mystring2,lista_2)).pack()



#Generación de caja de entrada.
e3 = tk.Entry(root,textvariable = mystring3,width=100,fg="blue",bd=3,selectbackground='Violet').pack()

#Generación de boton
button4 = tk.Button(root, 
                text='LIGA DE TABLA AUXILIAR', 
                fg='White', 
                bg= 'dark green',height = 1, width = 20,command=partial(getvalue,mystring3,lista_3)).pack()








#Generación de boton
button3 = tk.Button(root, 
                text='EJECUTAR', 
                fg='White', 
                bg= 'dark green',height = 1, width = 20,command=partial(funcion_principal,lista_1,lista_2,lista_3)).pack()


#loop principal corriendo
root.mainloop()
